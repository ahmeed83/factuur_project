import sys
from datetime import date, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font
from pathlib import Path
from jproperties import Properties
from excel2pdf import convert_excel_to_pdf
import os

# Read properties from property file
configs = Properties()
with open('app-config.properties', 'rb') as config_file:
    configs.load(config_file)

# Load Excel file
wb = load_workbook('invoice_template.xlsx')

ws = wb.active

date_format = '%d-%m-%Y'
today = date.today()
last_day_of_prev_month = today.replace(day=1) - timedelta(days=1)
start_day_of_prev_month = today.replace(day=1) - timedelta(days=last_day_of_prev_month.day)

datum_werkzaamheden = 'voor periode %s t/m %s' % \
                      (start_day_of_prev_month.strftime(date_format), last_day_of_prev_month.strftime(date_format))

hours = sys.argv[1]

ws['C26'] = int(hours)
ws['D26'] = int(configs.get("CLIENT_HOUR_RATE").data)
ws['D29'] = int(configs.get("CLIENT_RATE").data)
ws['B26'] = configs.get("CLIENT_DESCRIPTION").data
ws['B27'] = datum_werkzaamheden
ws['B29'] = configs.get("CLIENT_RATE_DESCRIPTION").data

ws['D2'] = configs.get("CLIENT_NAME").data
ws['D2'].font = Font(bold=True, size=20)
ws['D3'] = configs.get("CLIENT_ADDRESS_1").data
ws['D4'] = configs.get("CLIENT_ADDRESS_2").data

ws['D7'] = 'Factuur'
ws['D8'] = 'Factuurnummer: %s' % start_day_of_prev_month.strftime('%Y-%m')
ws['D9'] = 'Datum: %s' % today.strftime(date_format)

ws['B7'] = configs.get("COMPANY_NAME").data
ws['B7'].font = Font(bold=True, size=12)
ws['B8'] = configs.get("COMPANY_ADDRESS_1").data
ws['B9'] = configs.get("COMPANY_ADDRESS_2").data

ws['B12'] = 'Telefoon: ' + configs.get("COMPANY_TELEFOON").data
ws['B13'] = 'Email : ' + configs.get("COMPANY_EMAIL").data
ws['B15'] = 'K.v.k : ' + configs.get("COMPANY_KVK").data
ws['B16'] = 'BTW-nummer : ' + configs.get("COMPANY_BTW").data
ws['B17'] = 'IBAN : ' + configs.get("COMPANY_IBAN").data

home = str(Path.home())
fileName = '%s/Desktop/sample.xlsx' % home

wb.save(fileName)

convert_excel_to_pdf()

os.remove(fileName)
